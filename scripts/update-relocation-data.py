import datetime as dt
import html
import json
import os
import re
import sys
import urllib.request
import zipfile


HRSA_URL = (
    "https://bhw.hrsa.gov/data-research/projecting-health-workforce-supply-demand/"
    "technical-documentation/nursing"
)
CENSUS_T13_URLS = [
    "https://www2.census.gov/programs-surveys/demo/tables/geographic-mobility/2024/state-to-state-migration/State_to_State_Migration_Table_2024_T13.xlsx",
    "https://www2.census.gov/programs-surveys/demo/tables/geographic-mobility/2023/state-to-state-migration/State_to_State_Migration_Table_2023_T13.xlsx",
    "https://www2.census.gov/programs-surveys/demo/tables/geographic-mobility/2022/state-to-state-migration/State_to_State_Migration_Table_2022_T13.xlsx",
]

STATE_NAME_TO_ABBR = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "Puerto Rico": "PR",
}


def fetch_url(url, retries=2):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )
    import time
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return resp.read()
        except urllib.error.HTTPError as e:
            if attempt < retries and e.code in (403, 429, 503):
                wait = 5 * (attempt + 1)
                print(f"  HTTP {e.code} from {url}, retrying in {wait}s... (attempt {attempt + 1}/{retries + 1})")
                time.sleep(wait)
            else:
                raise


def normalize_text(raw):
    text = raw.decode("utf-8", "ignore")
    text = html.unescape(text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\bMaine a\b", "Maine", text)
    text = re.sub(r"(\d)\s+%", r"\1%", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def extract_hrsa_section(text):
    start = text.find("Exhibit IV-9: Distribution of Annual Nurse Migration to Destination State")
    if start == -1:
        start = text.find("Exhibit IV-9: Distribution of Annual Nurse Migration")
    if start == -1:
        start = text.find("Exhibit IV-9")
    if start == -1:
        raise RuntimeError("Could not find Exhibit IV-9 section.")
    end = text.find("U.S.", start)
    if end == -1:
        end = text.find("Exhibit IV-10", start)
    if end == -1:
        end = start + 20000
    return text[start:end + 200]


def parse_hrsa_migration(section):
    results = {}
    missing = []
    for state_name, abbr in STATE_NAME_TO_ABBR.items():
        pattern = (
            rf"{re.escape(state_name)}\s+([0-9,]+)\s+([0-9.]+%)\s+"
            rf"([0-9,]+)\s+([0-9.]+%)\s+([0-9,]+)\s+([0-9.]+%)"
        )
        match = re.search(pattern, section)
        if not match:
            missing.append(state_name)
            continue
        rn_bacc_annual = int(match.group(1).replace(",", ""))
        rn_bacc_dist = float(match.group(2).replace("%", ""))
        rn_ad_annual = int(match.group(3).replace(",", ""))
        rn_ad_dist = float(match.group(4).replace("%", ""))
        lpn_annual = int(match.group(5).replace(",", ""))
        lpn_dist = float(match.group(6).replace("%", ""))
        results[abbr] = {
            "lpn_annual": lpn_annual,
            "lpn_dist": lpn_dist,
            "rn_bacc_annual": rn_bacc_annual,
            "rn_bacc_dist": rn_bacc_dist,
            "rn_ad_annual": rn_ad_annual,
            "rn_ad_dist": rn_ad_dist,
        }
    if missing:
        print(f"Warning: missing HRSA rows for {missing}", file=sys.stderr)
    return results


def build_hrsa_shares(rows):
    totals = {
        "rn_total": 0,
        "lpn_total": 0,
        "rn_bacc_total": 0,
        "rn_ad_total": 0,
    }
    for data in rows.values():
        totals["rn_bacc_total"] += data["rn_bacc_annual"]
        totals["rn_ad_total"] += data["rn_ad_annual"]
        totals["lpn_total"] += data["lpn_annual"]
    totals["rn_total"] = totals["rn_bacc_total"] + totals["rn_ad_total"]

    rn_share = {}
    lpn_share = {}
    for abbr, data in rows.items():
        rn_annual = data["rn_bacc_annual"] + data["rn_ad_annual"]
        rn_share[abbr] = rn_annual / totals["rn_total"] if totals["rn_total"] else 0
        lpn_share[abbr] = data["lpn_annual"] / totals["lpn_total"] if totals["lpn_total"] else 0
    return rn_share, lpn_share


def try_load_census_migration(tmp_path):
    for url in CENSUS_T13_URLS:
        try:
            data = fetch_url(url)
        except Exception:
            continue
        with open(tmp_path, "wb") as f:
            f.write(data)
        if not zipfile.is_zipfile(tmp_path):
            continue
        return url
    return None


def compute_general_migration_share(tmp_path):
    try:
        from openpyxl import load_workbook
    except Exception:
        print("openpyxl not available; skipping Census migration share.", file=sys.stderr)
        return None

    if not zipfile.is_zipfile(tmp_path):
        return None

    wb = load_workbook(tmp_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for idx, row in enumerate(rows[:50]):
        if row and "State of current residence" in str(row[0] or ""):
            header_row_idx = idx
            break

    if header_row_idx is None:
        return None

    header = rows[header_row_idx]
    state_start_idx = None
    for i, cell in enumerate(header):
        if cell and "State of previous residence" in str(cell):
            state_start_idx = i + 1
            break

    if state_start_idx is None:
        return None

    state_names = []
    for cell in header[state_start_idx:]:
        if cell is None:
            break
        name = str(cell).strip()
        if name in STATE_NAME_TO_ABBR:
            state_names.append(name)
        else:
            state_names.append(name)

    inbound = {abbr: 0 for abbr in STATE_NAME_TO_ABBR.values()}
    for row in rows[header_row_idx + 1 :]:
        if not row or row[0] is None:
            continue
        current_state = str(row[0]).strip()
        if current_state not in STATE_NAME_TO_ABBR:
            continue
        abbr = STATE_NAME_TO_ABBR[current_state]
        for idx, origin_name in enumerate(state_names):
            if origin_name == current_state:
                continue
            val = row[state_start_idx + idx]
            if isinstance(val, (int, float)):
                inbound[abbr] += val

    total_inbound = sum(inbound.values())
    if not total_inbound:
        return None
    return {abbr: val / total_inbound for abbr, val in inbound.items()}


def minmax_scale(values):
    if not values:
        return {}
    min_v = min(values.values())
    max_v = max(values.values())
    if max_v == min_v:
        return {k: 50.0 for k in values}
    return {k: 100 * (v - min_v) / (max_v - min_v) for k, v in values.items()}


def main():
    output_dir = os.environ.get("LNI_OUTPUT_DIR", "public/data")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "relocation.json")

    now = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    refresh_days = 7

    if os.path.exists(output_path):
        try:
            with open(output_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
            last = existing.get("lastUpdated")
            if last:
                last_dt = dt.datetime.fromisoformat(last.replace("Z", "+00:00"))
                if (dt.datetime.now(dt.timezone.utc) - last_dt).days < refresh_days:
                    print("Relocation data is fresh; skipping refresh.")
                    return
        except Exception:
            pass

    try:
        hrsa_html = fetch_url(HRSA_URL)
        section = extract_hrsa_section(normalize_text(hrsa_html))
        hrsa_rows = parse_hrsa_migration(section)
        rn_share, lpn_share = build_hrsa_shares(hrsa_rows)
    except Exception as e:
        print(f"Warning: Could not fetch HRSA data ({e}). Checking for existing data...")
        if os.path.exists(output_path):
            print(f"  Using existing {output_path} — skipping update.")
            return
        print("  No existing data found. Generating with empty HRSA data.")
        rn_share, lpn_share = {}, {}

    tmp_census = os.path.join(output_dir, "_tmp_census.xlsx")
    census_url = try_load_census_migration(tmp_census)
    general_share = compute_general_migration_share(tmp_census) if census_url else None

    relocation_source = {}
    combined_share = {}
    for abbr in rn_share:
        if rn_share.get(abbr):
            combined_share[abbr] = rn_share[abbr]
            relocation_source[abbr] = "rn"
        elif lpn_share.get(abbr):
            combined_share[abbr] = lpn_share[abbr]
            relocation_source[abbr] = "clinical"
        elif general_share and general_share.get(abbr):
            combined_share[abbr] = general_share[abbr]
            relocation_source[abbr] = "general"
        else:
            combined_share[abbr] = 0
            relocation_source[abbr] = "unknown"

    # Blend RN/clinical with general migration if available
    blended_share = {}
    for abbr in combined_share:
        base = combined_share[abbr]
        if general_share and abbr in general_share and relocation_source[abbr] in ("rn", "clinical"):
            blended_share[abbr] = 0.7 * base + 0.3 * general_share[abbr]
        else:
            blended_share[abbr] = base

    relocation_scale = minmax_scale(blended_share)

    payload = {
        "lastUpdated": now,
        "refreshDays": refresh_days,
        "sources": {
            "rnMigration": HRSA_URL,
            "generalMigration": census_url,
        },
        "weights": {
            "rn": 0.7,
            "general": 0.3,
            "fallbackClinical": "lpn",
            "fallbackGeneral": "census"
        },
        "rnDestinationShare": rn_share,
        "clinicalDestinationShare": lpn_share,
        "generalMigrationShare": general_share or {},
        "relocationScale": relocation_scale,
        "relocationSource": relocation_source,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, sort_keys=True)
    print(f"Wrote relocation data to {output_path}")


if __name__ == "__main__":
    main()
